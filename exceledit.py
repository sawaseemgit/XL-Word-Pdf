import openpyxl as xl
import os

wrbk = xl.Workbook()  # Creates New/Blank Excel sheet
sheets = wrbk.sheetnames  # gets all sheets names
sheet = wrbk[sheets[0]]

# for i in range(1, 3):
#     print(sheet.cell(row=i, column=2).value)
os.chdir('F:\\Python\\XL-Word-PDF')
sheet2= wrbk.create_sheet(index=0,title='NewSheet')
# Sheet2.title  # To change sheet title
sheet2['A2'] = 'Apples'
sheet2['B2'] = '22'
wrbk.save('writeExfile2.xlsx')  # creates & saves in new Excel file


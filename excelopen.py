import openpyxl
import os

os.chdir('F:\\Python\\XL-Word-PDF\\automate_online-materials')

wrkbook = openpyxl.load_workbook('example.xlsx')
sheet = wrkbook['Sheet1']
# print(str(sheet['A1'].value)) #another way for Cell obj in line10

print(str(sheet.cell(row=2, column=3)))  # returns Cell object
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
